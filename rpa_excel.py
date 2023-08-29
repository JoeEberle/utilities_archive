from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, GradientFill, Alignment, Color, Fill
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import colors
from openpyxl.styles import numbers 
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows 
from openpyxl.formatting.rule import DataBarRule
from openpyxl.formatting.rule import ColorScaleRule
import openpyxl
import glob 
import configparser   
from datetime import date, datetime
import logging # built in python library that does not need to be installed 
import quick_logger as ql

LIGHT_BLUE = "C5D9F1"
YELLOW = "FFFF00"
GRAY = "E4DFEC"
DEEP_BLUE = "1072BA"
WHITE = "FFFFFF"

CONTENT_TAB_COLOR = DEEP_BLUE
USER_TAB_COLOR = YELLOW
BLANK_WHITE = WHITE 
HEADER_COLOR = LIGHT_BLUE  # Light Blue - Light Blue indicates relevant content for validation 
HEADER_COLOR_USER_ENTRY = YELLOW # Yellow - Yellow indicates a field for User Entry 
HEADER_COLOR_INFORMATIONAL_ONLY = GRAY # Gray - Indicates a field that maybe helpful for information purposes   

Green_Good  = "85F386"
Purple_Perfect = "E951F6"
Yellow_OK = "F7F173" 
Gray_None = "696968"  
Red_Bad = "E8241B" 

Yellow_OK_Fill = PatternFill(start_color = Yellow_OK, end_color = Yellow_OK, fill_type = "solid")
Purple_Perfect_Fill = PatternFill(start_color = Purple_Perfect, end_color = Purple_Perfect, fill_type = "solid")
Green_Good_Fill = PatternFill(start_color = Green_Good, end_color = Green_Good, fill_type = "solid")
Red_Bad_Fill = PatternFill(start_color = Red_Bad, end_color = Red_Bad, fill_type = "solid")
Gray_None_Fill = PatternFill(start_color = Gray_None, end_color = Gray_None, fill_type = "solid")
 
# establish the configuration settings parser
config = configparser.ConfigParser()  
cfg = config.read('config.ini')   

def get_column_width(worksheet,column,row, default_width):
    status = f'\n Worksheet column:{column} and row:{row} ' 
    value = worksheet[get_column_letter(column) + str(row)].value
    if value == None:
        length_of_column = default_width  # if the column sample contains nothing set width = 10 
    if value != None:
        if type(value) == int or type(value) == float:  
            length_of_column = default_width # if the column sample numeric set width = 10
        elif type(value) == 'datetime.datetime':
            length_of_column = default_width  
        elif type(value) == 'datetime64':
            length_of_column = default_width  
        elif type(value) == 'float64':
            length_of_column = default_width     
        elif type(value) == object:
            length_of_column = len(value) + 25 # Add a buffer of 25 characters for strings             
        else:         
            length_of_column = 30
        status = status + f'\nValue of column ={value} with length:{length_of_column}'
#     print(status)
    return length_of_column  

def get_column_type(worksheet,column,row):
    value = worksheet[get_column_letter(column) + str(row)].value
    type_of_column = type(value)
    return type_of_column  

def get_meta_tag():
    from datetime import datetime
    date_time = datetime.now() # Returns a datetime object containing the local date and time
    date_stamp = f'{date_time.year}-{date_time.month}-{date_time.day}' 
    time_stamp = f'{date_time.hour}:{date_time.minute}:{date_time.second}.{date_time.microsecond} '
    meta_tag = f'Formatted on:{date_stamp} at {time_stamp}'   
    return meta_tag 

def format_column_header(ws):
    ''' Format the Header as text bold and backgound light blue centered wordwrap ''' 
    print(f'format_column_header max_row:{ws.max_row} max_col:{ws.max_column} ')
    if ws.max_row > 3: # Dont bother with header formatting anything less than 3 rows 
        for col in range(1,ws.max_column + 1):
            header_value = str(ws[get_column_letter(col) + '2'].value)
#             print(f'format_column_header value:{header_value}')
            if len(header_value) >= 1:  # Only format headers if greater than 1 letters 
                ws[get_column_letter(col) + '2'].font = Font(bold=True ) 
                ws[get_column_letter(col) + '2'].fill = PatternFill("solid", start_color=HEADER_COLOR) 
                ws[get_column_letter(col) + '2'].alignment = Alignment(horizontal='center', wrap_text=True)
                ws[get_column_letter(col) + '2'].value = str(ws[get_column_letter(col) + '2'].value).title().replace('_',' ')  
    return ws.max_column       


def color_column(ws, column):
    ''' Color code the column based upon its value ''' 
    
    column = 0 
    for col in range(1,ws.max_column + 1):
        header_value = str(ws[get_column_letter(col) + '2'].value)  
#         print(f'checking header value {header_value} to determine if coloring is necessary {col}')
        
        if header_value == 'Distance To Goal %':
            column = col  
#             print(f'Found {header_value} column {column} to color in position {col}')  
     
    if column == 0:
#         print(f'Found no column {column} to color')        
        return      #   Not Coloring sheet - Returning column greater than 8

#     print(f'About to color column:{column}')         
    for row in range(3,ws.max_row + 1):
        column_value = ws[get_column_letter(column) + str(row)].value
#         print (f'Column value = {column_value} type {type(column_value)}')
#         if type(column_value) == 'float' or type(column_value) == 'int': 
        if column_value < 0:
            ws[get_column_letter(column) + str(row)].fill = Purple_Perfect_Fill
#             print (f'Color = Purple_Perfect_Fill')
        elif ((column_value >= 0) and (column_value < 10)):
            ws[get_column_letter(column) + str(row)].fill = Green_Good_Fill    
#             print (f'Color = Green_Good_Fill')                
        elif ((column_value >= 10) and (column_value < 15)):
            ws[get_column_letter(column) + str(row)].fill = Yellow_OK_Fill   
#             print (f'Color = Yellow_OK_Fill')                    
        elif ((column_value >= 15) and (column_value < 99)):
            ws[get_column_letter(column) + str(row)].fill = Red_Bad_Fill  
#             print (f'Color = Red_Bad_Fill')                    
        elif (column_value == 0) :
            ws[get_column_letter(column) + str(row)].fill = Gray_None_Fill   
#             print (f'Color = Gray_None_Fill')                       
        else:
            ws[get_column_letter(column) + str(row)].fill = Red_Bad_Fill 
#             print (f'Color = Red_Bad_Fill')                      
   
    return ws.max_column  

def get_author():
    meta_tag = 'Joe Eberle'    
    return meta_tag 

def add_logo_image(current_ws):
    ''' The add_logo_image functon adds a logo to the current worksheet'''
    img = Image( config.get('global_infrastructure', 'logo_file'))
 
    if current_ws.max_column  > 6:
        img.anchor = 'F1' 
    else:
        img.anchor = 'D1'      
    img.width = 180
    img.height = 30
    img.alignment = Alignment(horizontal='center', vertical='center')          
    current_ws.add_image(img) 
    return 'logo image added'

def align_numerical_columns(current_ws):
    ''' The align_numerical_columns functon aligns all numeric columns in the current worksheet'''
    max_cols = current_ws.max_column + 1
    for col in range(1,max_cols):
        column_type = get_column_type(current_ws,col,3)
        if column_type == int:
            current_ws.column_dimensions[get_column_letter(col)].alignment = Alignment(horizontal='right')            
        elif column_type == float: 
            current_ws.column_dimensions[get_column_letter(col)].alignment = Alignment(horizontal='right')            
        else:
            current_ws.column_dimensions[get_column_letter(col)].alignment = Alignment(horizontal='left')  
    return 'numeric columns aligned'      


def create_report_title(current_ws, template, filename):
    import os
    ''' The create_report_title adds a title and format to the current sheet.'''
    # Make additional formatting Changes if a template other than "standard" is used. 
    if template == 'report' or  template == 'standard':
    # Create The Title 
        current_ws.insert_rows(1)
        current_ws.merge_cells('A1:' + get_column_letter(current_ws.max_column) + '1')
        current_ws.row_dimensions[1].height = 30 # Large enough to support title and  Logo image
        top_left_cell = current_ws['A1']
        report_header = os.path.basename(filename)   # Remove the path from the filename
        report_header = os.path.splitext(report_header)[0] # Remove the extension 
        header = report_header.replace('_',' ').title() # use Title to uppercase each word in the header
        header = header.replace('Pcp','PCP').replace('Gppc','GPPC') # custom uppercase for specific practices
        top_left_cell.value = header 
 
    # Gradient Color Fill the Title in Row 1   
        top_left_cell.fill = PatternFill("solid", fgColor=CONTENT_TAB_COLOR)
        top_left_cell.fill = GradientFill(stop=(CONTENT_TAB_COLOR, BLANK_WHITE))
        top_left_cell.font = Font(b=True, color="000000")
        top_left_cell.alignment = Alignment(horizontal="left", vertical="center")
        fontStyle = Font(size = "18", bold = "True")
        top_left_cell.font = fontStyle  
    return 'Report title Created'

def add_meta_tab(work_book, report_header, filename):
    ''' Add a tab to the workbook that contains meta data about the workbook.'''
    import os
  
    ws = work_book.create_sheet('Workbook Info')
    ws.sheet_properties.tabColor = USER_TAB_COLOR
     
# Work Sheet 4 - Patient Attribution - Set Column Widths and Header Hieght
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 10    

# Insert rows for spreadsheet title and image
    fontStyle = Font(size = "12", bold = True)
    ws.insert_rows(1,1)        
    
# Insert Data Services Logo    
    img = Image( config.get('global_infrastructure', 'logo_file'))
    img.anchor = 'B1' 
    img.width = 180
    img.height = 30
    img.alignment = Alignment(horizontal="center" , vertical="center")
    ws.add_image(img)   
    
# Create The Title 
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 40
    top_left_cell = ws['A1']
    top_left_cell.value = "Workbook Info"       
    
# Show last updated date and report description 
    from datetime import datetime
    import time 
    date_time = datetime.now() # Returns a datetime object containing the local date and time
    date_stamp = f'{date_time.strftime("%Y-%m-%d %H:%M:%S")}' 


    ws['A2'].value = 'Formatted and Published on' 
    ws['B2'].value = f'{date_stamp}'    
    ws['A3'].value = 'Original Data File' 
    ws['B3'].value = f'{filename}' 
    ws['A4'].value = 'Original Data File Size' 
    ws['B4'].value = f'{os.path.getsize(filename)}'     
    ws['A5'].value = 'Original Data File Create Date' 
    ws['B5'].value = f'{time.ctime(os.path.getctime(filename))}'   
    ws['A6'].value = 'Original Data File Modified Date' 
    ws['B6'].value = f'{time.ctime(os.path.getmtime(filename))}'    
#     ws['A7'].value = 'Department' 
#     ws['B7'].value = f'Kaleida IT Data Services Team - Dave Barnes'
#     ws['A8'].value = 'Team' 
#     ws['B8'].value = f'Gina Mucha, Helmi (Al) Seoud, Glenn Thomakos, Janet Crist, Frank Metty, Anna Mucha, Joe Eberle'
  
    top_left_cell.fill = PatternFill("solid", fgColor=CONTENT_TAB_COLOR)
    top_left_cell.fill = GradientFill(stop=(USER_TAB_COLOR, BLANK_WHITE))
    top_left_cell.font  = Font(b=True, color="000000")
    top_left_cell.alignment = Alignment(horizontal="left", vertical="center")
    fontStyle = Font(size = "18", bold = "True")
    top_left_cell.font = fontStyle 
    return work_book

   
def list_sheets(wb):
    ''' The list_sheets function lists all the sheets in a workbook. '''
    return wb.sheetnames  

    
def format_worksheet(ws, filename, template='standard'):
    ''' The format_worksheet formats all worksheets with the same style, look and feel. '''
    ws.sheet_properties.tabColor = CONTENT_TAB_COLOR 
    max_cols = ws.max_column + 1   
    status = create_report_title(ws, 'report', filename) # format fancy title 
    
    debugging = False
    if debugging:
        print(f'\n\nFormatting file:{filename} with maximum Columns:{max_cols} template:{template} ')
    
    status = add_logo_image(ws) # Insert worksheet Logo 
 
    if debugging: 
        print(f'after add logo ')
 
    if ws.max_row > 5: # Set the range of data to filter - Dont bother filtering anything less than 3 rows 
        full_data_range = "A2:" + get_column_letter(ws.max_column) + str(ws.max_row)
        ws.auto_filter.ref = full_data_range         
 
    format_column_header(ws)     # Format the Header as text bold and backgound light blue centered wordwrap
    
    
    default_column_width = 15      
    if template == 'quality_report':
       ws.row_dimensions[2].height = 43 # Make Large enough to stack three header lines
       default_column_width = 9  
        
    if template == 'standard':
       ws.row_dimensions[2].height = 43 # Make Large enough to stack three header lines
       default_column_width = 9          
        

    for col in range(1,max_cols): # Set Column Widths using function 
        ws.column_dimensions[get_column_letter(col)].width = get_column_width(ws,col,3,default_column_width)
 
    status = align_numerical_columns(ws) # Set Column alignment using function  
 
    for col in range(1,(max_cols)): # Loop through the rows to set the desired number format 
        column_header =  ws[get_column_letter(col) + '2'].value.upper() 
        if debugging: 
            print(f'formatting number column type with header:{column_header}')
        number_fmt = '#,##0_);[Red](#,##0)'
        if column_header.find('DATE') > -1:   
             number_fmt = numbers.FORMAT_DATE_DDMMYY 
        if column_header.find('DOB') > -1:   
             number_fmt = numbers.FORMAT_DATE_DDMMYY                  
        if column_header.find('PHONE') > -1:   
             number_fmt = numbers.FORMAT_TEXT      
        if column_header.find('NPI') > -1:   
             number_fmt = numbers.FORMAT_TEXT          
        if column_header.find('CLAIM') > -1:   
             number_fmt = numbers.FORMAT_TEXT                      
        for row in range(2,ws.max_row + 1):
            ws[get_column_letter(col) + str(row)].number_format = number_fmt      
            
    if debugging: 
        print(f'Formatting file:{filename} with maximum Coluns:{max_cols} template:{template} ')            
            
    if template == 'quality_report':
        status = color_column(ws, 8) # color code the performance of the measure
 
    full_data_range = "A2:" + get_column_letter(ws.max_column) + str(ws.max_row)
    ws.auto_filter.ref = full_data_range 
 
    ws.freeze_panes = 'A3'   # Freeze so the header columns are always on screen 
 
    from openpyxl.comments import Comment # Add Meta tagging - information about information 
#     ws['A2'].comment = Comment(get_meta_tag(),get_author())  
    return f'Worksheet {ws.title}'

# Set the TAB and color it  
def format_workbook(filename, template='standard'):
    ''' The format_workbook performs remote process automation to format an excel workbook.'''
    status = f'\nRead in raw file:{filename}'
    wb = load_workbook(filename)
    format_filename = filename.replace('.xlsx','_ft.xlsx') 
    
    for sheet in wb:
        status =  status + '\n'+ format_worksheet(sheet, filename, template)
        
# Save the file to persist the changes  
    add_meta_tab(wb, 'report_header', filename) 
    wb.save(format_filename)
    status = status + f'\nFormatted:{format_filename}'
    return status      


def format_excel_directory(directory, template='standard'):
    formatted_files = 0
    files = glob.glob(directory + '\\*.xlsx')
    status = f'Found {len(files)} excel files in directory:{directory}'
    print(status)
    logging.info(f'Found {len(files)} excel files in directory:{directory}')        
    for file in files:
        if (file.find('_ft') == -1) & (file.find('_pb') == -1) & (file.find('~$') == -1): # already formatted 
            print(f'File name to be loaded:{file}') 
            status = format_workbook(file, template)  
            formatted_files += 1 
            logging.info(f'{status}')                   
            print(status) 
    return formatted_files   

def format_excel_file(full_file_path, template='standard'):
    formatted_files = 0
    logging.info(f'Formatting {full_file_path} excel file.')        
    status = format_workbook(full_file_path, template)  
    return status   